"""
App keywords that are executed for Check in out tasks. 
"""
import openpyxl
from os.path import exists
from datetime import datetime
from taskslocales import *
from taskslocales import _
from common_keywords import *

class CustomKeywords:
    def __init__(self):
        self.workbook = None
        self.sheet = None
        self.common = CommonKeywords()

    def check_in_app_task(self):
        """Check in App task"""
        self.open_checkin_app()

        env = self.common.load_vault_file()
        if env['LEVEL_3_ACTIONS']['DO_CHECKIN_ACTION']:
            # Find first empty row and fill current checkin date and time
            next_row = self.find_empty_row()
            date_now = datetime.now().strftime('%Y-%m-%d %H:%M')
            self.sheet.cell(row=next_row, column=1, value=date_now)
            self.workbook.save(env['MY_DATA']['CHECKIN']['XLS'])
        else:
            self.common.pause_execution(
                _('Please record the time and then press OK to continue')
            )
        self.close_checkin_app()

    def check_out_app_task(self):
        """Check out App task"""
        self.open_checkin_app()

        env = self.common.load_vault_file()
        if env['LEVEL_3_ACTIONS']['DO_CHECKOUT_ACTION']:
            next_row = self.find_empty_row()
            date_now = datetime.now().strftime('%Y-%m-%d %H:%M')
            self.sheet.cell(row=next_row-1, column=2, value=date_now)
            self.workbook.save(env['MY_DATA']['CHECKIN']['XLS'])
        else:
            self.common.pause_execution(
                _('Please record the time and then press OK to continue')
            )
        self.close_checkin_app()

    def verify_app_task(self):
        """Verify App task"""
        env = self.common.load_vault_file()
        if os.name == 'nt':
            os.startfile(env['MY_DATA']['CHECKIN']['XLS'])
        else:
            subprocess.call(["open", env['MY_DATA']['CHECKIN']['XLS']])

    def open_checkin_app(self):
        """Open Excel file"""
        env = self.common.load_vault_file()

        # create file if it does not exist
        if not exists(env['MY_DATA']['CHECKIN']['XLS']):
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.workbook.save(env['MY_DATA']['CHECKIN']['XLS'])
            self.workbook.close()

        # When level 3 is enabled, edit excel directly without opening it, otherwise open it with system default application
        if env['LEVEL_3_ACTIONS']['DO_CHECKIN_ACTION']:
            self.workbook = openpyxl.load_workbook(env['MY_DATA']['CHECKIN']['XLS'])
            self.sheet = self.workbook.active
        else:
            if os.name == 'nt':
                os.startfile(env['MY_DATA']['CHECKIN']['XLS'])
            else:
                subprocess.call(["open", env['MY_DATA']['CHECKIN']['XLS']])

    def close_checkin_app(self):
        """Close the browser or Excel file"""
        env = self.common.load_vault_file()
        if env['LEVEL_3_ACTIONS']['DO_CHECKIN_ACTION']:
            self.workbook.close()

    def custom_app_task(self):
        """Custom App task"""
        self.open_custom_app()
        env = self.common.load_vault_file()
        
        if env['LEVEL_3_ACTIONS']['DO_CUSTOM_ACTION']:
            self.common.pause_execution(_('Custom task action was not implemented!'))
        else:
            self.common.pause_execution(_('Custom task message action was not defined.'))

        self.close_custom_app()

    def open_custom_app(self):
        """Open custom URL or file"""
        # Not implemented in the Excel version
        pass

    def close_custom_app(self):
        """Close custom URL or file"""
        # Not implemented in the Excel version
        pass
        
    def find_empty_row(self):
        """Find the first empty row in the Excel sheet"""
        row = 1
        while self.sheet.cell(row=row, column=1).value is not None:
            row += 1
        return row